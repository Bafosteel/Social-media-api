import requests
import hashlib
import json
from requests_toolbelt.multipart.encoder import MultipartEncoder
from tkinter import *
from tkinter.filedialog import askopenfilename
import openpyxl
import ctypes

Token='TOKEN'
session_secret_key='SECRET_KEY'
byte_session = bytes(str(session_secret_key),"utf8")
application = 'PUBLIC_KEY'
byte_application=bytes(str(application),"utf8")
group_id = str('GROUP_ID')
byte_group_id = bytes(str(group_id), "utf8")

# Class for VK

class Vk():

    def get_videos(self,path_to_file,name_of_video, token, group_id_1):
        ids = []
        for i in range(0, len(path_to_file)):
            try:
                vid = {'video': (str(path_to_file[i]),
                                 open(r''+str(path_to_file[i]), 'rb'))}
                method = 'https://api.vk.com/method/video.save?'
                data = dict(access_token=token, gid=group_id_1,name=name_of_video[i])
                response = requests.post(method, data)
                result = json.loads(response.text)

                upload_url = result['response']['upload_url']
                response = requests.post(upload_url, files=vid)
                result = json.loads(response.text)

                ids.append('video-' + str(group_id_1) + '_' + str(result['video_id']))
            except:
                # print("This is something else")
                ids.append(str(path_to_file[i]))
        return ids

    def get_image_ids(self,path_to_file,token,group_id_1):
        # Получаем id изображения
        id_s=[]
        for path in path_to_file:
            try:
                # путь к вашему изображению
                img = {'photo': (str(path), open(r''+str(path), 'rb'))}

                # Получаем ссылку для загрузки изображений
                method_url = 'https://api.vk.com/method/photos.getWallUploadServer?'
                data = dict(access_token=token, gid=group_id_1)
                response = requests.post(method_url, data)
                result = json.loads(response.text)
                upload_url = result['response']['upload_url']

                # Загружаем изображение на url
                response = requests.post(upload_url, files=img)
                result = json.loads(response.text)

                # Сохраняем фото на сервере и получаем id
                method_url = 'https://api.vk.com/method/photos.saveWallPhoto?'
                data = dict(access_token=token, gid=group_id_1, photo=result['photo'], hash=result['hash'],
                            server=result['server'])
                response = requests.post(method_url, data)
                result = json.loads(response.text)['response'][0]['id']
                id_s.append(result)
            except:
                print('This is href')
                id_s.append(str(path))

        return id_s

    def start_Vk(self,private_post,post_with_button, message,type_of_attachment,
                 path_to_file,link_button,link_title,link_image,name_of_video, Token, group_ids):
        result = self.get_image_ids(path_to_file,Token,group_ids)
        video = self.get_videos(path_to_file,name_of_video,Token,group_ids)
        posts = []
        group_id = "-" + str(group_ids)
        buttons = {'Запустить':'app_join','Играть':'app_game_join','Перейти':'open_url','Открыть':'open',
                   'Подробнее':'more','Позвонить':'call','Забронировать':'book','Записаться':'enroll',
                   'Зарегистрироваться':'register','Купить':'buy','Купить билет':'but_ticket','Заказать':'order',
                   'Установить':'install','Связаться':'contact','Заполнить':'fill','Подписаться':'join_public',
                   'Я пойду':'join_event','Вступить':'join','Связаться_2':'im','Написать':'im2'}

        # Создаем посты
        for i in range(0,len(private_post)):
            if ('photo' in type_of_attachment[i]):
                # print('photo')
                if ("private" not in private_post[i]):
                    # print("public")
                    # Создаем пост
                    r = requests.post('https://api.vk.com/method/wall.post',
                                    params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                              'message': str(message[i]),
                                              'attachments': str(result[i])})
                    response = r.json()
                    posts.append('https://vk.com/wall-'+str(group_id)[1:]+'_'+str(response['response']['post_id']))
                else:
                    # print("private")
                    if ('yes' in post_with_button[i]):
                        # Создаем скрытый пост с кнопкой и с фото
                        r = requests.post('https://api.vk.com/method/wall.postAdsStealth',
                                        params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                                'message': str(message[i]),
                                                'link_button': str(buttons[link_button[i]]),
                                                'link_image': str(link_image[i]),
                                                'link_title': str(link_title[i]),
                                                'attachments': str(result[i])})
                        response = r.json()
                        posts.append('https://vk.com/wall-' + str(group_id)[1:]+'_'+ str(response['response']['post_id']))
                    else:
                        # Создаем скрытый пост без кнопки и с фото
                        r = requests.post('https://api.vk.com/method/wall.postAdsStealth',
                                          params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                                  'message': str(message[i]),
                                                  'attachments': str(result[i])})
                        response = r.json()
                        posts.append(
                            'https://vk.com/wall-' + str(group_id)[1:] + '_' + str(response['response']['post_id']))
            else:
                # print('video')
                if ("private" not in private_post[i]):
                    # print("public")
                    # Создаем пост
                    r = requests.post('https://api.vk.com/method/wall.post',
                                    params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                              'message': str(message[i]),
                                              'attachments': str(video[i])})
                    response = r.json()
                    posts.append('https://vk.com/wall-'+str(group_id)[1:]+'_'+str(response['response']['post_id']))
                else:
                    # print("private")
                    if ('yes' in post_with_button[i]):
                        # Создаем скрытый пост с кнопкой и с видеозаписью
                        r = requests.post('https://api.vk.com/method/wall.postAdsStealth',
                                        params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                                'message': str(message[i]),
                                                'link_button': str(link_button[i]),
                                                'link_image': str(link_image[i]),
                                                'link_title': str(link_title[i]),
                                                'attachments': str(video[i])})
                        response = r.json()
                        posts.append('https://vk.com/wall-' + str(group_id)[1:]+'_'+ str(response['response']['post_id']))
                    else:
                        # Создаем скрытый пост без кноки, но с видеозаписью
                        r = requests.post('https://api.vk.com/method/wall.postAdsStealth',
                                          params={'owner_id': group_id, 'access_token': Token, 'from_group': 1,
                                                  'message': str(message[i]),
                                                  'attachments': str(video[i])})
                        response = r.json()
                        posts.append(
                            'https://vk.com/wall-' + str(group_id)[1:] + '_' + str(response['response']['post_id']))
        return posts

# Class for OK

class Ok():
    def get_video(self, path_to_file, name_of_video):
        print(path_to_file)
        byte_path_to_file = []
        byte_name_of_video = []
        errors_video = []
        for files in path_to_file:
            byte_path_to_file.append(bytes(str(files), "utf8"))
        for name in name_of_video:
            byte_name_of_video.append(bytes(str(name), "utf8"))
        ids = []

        for i in range(0, len(path_to_file)):
            try:
                video = dict(application_key=str(application), file_name=str(path_to_file[i]),
                             file_size=0, gid=str(group_id),
                             method='video.getUploadUrl', access_token=Token)

                signature_for_video = hashlib.md5(b"application_key=" + byte_application +
                                                  b"file_name=" + byte_path_to_file[i] +
                                                  b"file_size=0gid=" + byte_group_id +
                                                  b"method=video.getUploadUrl" + byte_session).hexdigest()

                video['sig'] = str(signature_for_video)
                url = 'https://api.ok.ru/fb.do'

                response = requests.post(url, video)

                video_id = str(response.json()['video_id'])
                url = response.json()['upload_url']
                print(video_id)
                print(url)
                print(response.json())

                file = open(path_to_file[i], 'rb')
                # 2. Загружаем ролик request.post
                multipart_data = MultipartEncoder(
                    fields={'video': ("sample", file, 'video/mp4')}
                )
                try:
                    response = requests.post(
                        url,
                        data=multipart_data,
                        headers={'Content-Type': multipart_data.content_type}
                    )
                    print(response.status_code)
                except Exception as ex:
                    print(u"1" + str(ex.message))
                    errors_video.append(ex.message)

                # Заканчиваем процесс заливки видео
                bytes_video_id = bytes(str(video_id), "utf8")
                # Зададим имя видеоролику с помощью параметра title
                video_update = dict(application_key=application,
                                    method='video.update',
                                    title=str(name_of_video[i]),
                                    vid=video_id, access_token=Token)

                signature_for_video_update = hashlib.md5(b"application_key=" + byte_application +
                                                         b"method=video.update"
                                                         b"title=" + byte_name_of_video[i] +
                                                         b"vid=" + bytes_video_id + byte_session).hexdigest()

                video_update['sig'] = str(signature_for_video_update)
                url = 'https://api.ok.ru/fb.do'
                response = requests.post(url, video_update)
                print(response.status_code)
                ids.append(str(video_id))
            except:
                print("This is something else")
                ids.append(str(path_to_file[i]))
        return ids

    def get_image_id(self, path_to_file):
        # Получаем id изображения
        id_s = []
        errors_photo = []
        print(path_to_file)
        for path in path_to_file:
            try:
                photo = dict(application_key=application, gid=group_id,
                             method='photosV2.getUploadUrl', access_token=Token)

                signature_for_photo = hashlib.md5(b"application_key=" + byte_application +
                                                  b"gid=" + byte_group_id +
                                                  b"method=photosV2.getUploadUrl" + byte_session).hexdigest()

                photo['sig'] = str(signature_for_photo)
                url = 'https://api.ok.ru/fb.do'

                response = requests.post(url, photo)

                # грузим фото на сервак
                url = response.json()['upload_url']
                photo_id = str(response.json()['photo_ids'][0])

                # Путь к изображению
                file = open(str(path), 'rb')

                multipart_data = MultipartEncoder(
                    fields={'photo': ("sample", file, 'photo/png')}
                )
                try:
                    response = requests.post(
                        url,
                        data=multipart_data,
                        headers={'Content-Type': multipart_data.content_type}
                    )
                    # Получим токен фотографии
                    photo_token = str(response.json()['photos'][photo_id]['token'])
                    id_s.append(photo_token)
                except Exception as ex:
                    print(u"1" + str(ex.message))
                    errors_photo.append(ex.message)
                print(id_s)
            except:
                print('This is href')
                id_s.append(str(path))

        return id_s

    def start_ok(self,private_post, post_with_button, message, type_of_attachment,
                 path_to_file, link_button, link_title, link_image, name_of_video):
        result = self.get_image_id(path_to_file)
        video = self.get_video(path_to_file, name_of_video)
        posts = []
        for i in range(0, len(private_post)):
            if ('photo' in type_of_attachment[i]):
                print('photo')
                if 'private' not in private_post[i]:
                    print('public')
                    print(result[i])
                    attachment = " {\"media\": [{\"type\": \"photo\",\"list\": " \
                                 "[{\"id\": \"" + result[i] + "\" }]},"" \
                                 ""{\"type\": \"link\",\"url\": \"https://apiok.ru/\"},"" \
                                 ""{\"type\": \"text\",\"text\": \"" + message[i] + "\"}]}"

                    byte_attachment = bytes(attachment, "utf8")

                    data = dict(application_key=application, attachment=attachment, format='json', gid=group_id,
                                method='mediatopic.post', type='GROUP_THEME', access_token=Token)

                    signature_for_post = hashlib.md5(b"application_key=" + byte_application + b"attachment="+
                                                     byte_attachment + b"format=jsongid=" + byte_group_id +
                                                     b"method=mediatopic.post"
                                                     b"type=GROUP_THEME" + byte_session).hexdigest()

                else:
                    print('private')
                    attachment = " {\"media\": [{\"type\": \"link\","" \
                    ""\"url\": \"" + result[i] + "\"}," "{\"type\": \"text\",\"text\": \"" \
                                     + message[i] + "\"}]}"

                    byte_attachment = bytes(attachment, "utf8")
                    data = dict(application_key=application, attachment=attachment, format='json', gid=group_id,
                                hidden_post='true',
                                method='mediatopic.post', type='GROUP_THEME', access_token=Token)

                    signature_for_post = hashlib.md5(
                        b"application_key=" + byte_application + b"attachment=" + byte_attachment
                        + b"format=jsongid=" + byte_group_id +
                        b"hidden_post=truemethod=mediatopic.post"
                        b"type=GROUP_THEME" + byte_session).hexdigest()

                data['sig'] = str(signature_for_post)
                url = 'https://api.ok.ru/fb.do'

                response = requests.post(url, data)
                print(response.json())
                if 'error_code' in response.json():
                    posts.append('None')
                else:
                    posts.append('https://ok.ru/group/' + str(group_id) + '/topic/' + str(response.json()))

            else:
                print('video')
                attachment = " {\"media\": [{\"type\": \"movie-reshare\"," \
                                "\"movieId\": \"" + str(video[i]) + "\"},"" \
                                ""{\"type\": \"link\",\"url\": "" \
                                ""\"https://apiok.ru/\"},"" \
                                ""{\"type\": \"text\",\"text\": \"" + message[i] + "\"}]}"
                byte_attachment = bytes(attachment, "utf8")

                if 'private' not in private_post[i]:
                    data = dict(application_key=application, attachment=attachment, format='json', gid=group_id,
                                method='mediatopic.post', type='GROUP_THEME', access_token=Token)

                    signature_for_post = hashlib.md5(
                        b"application_key=" + byte_application + b"attachment=" + byte_attachment
                        + b"format=jsongid=" + byte_group_id +
                        b"method=mediatopic.posttype=GROUP_THEME" + byte_session).hexdigest()

                else:
                    data = dict(application_key=application, attachment=attachment, format='json', gid=group_id,
                                hidden_post='true',
                                method='mediatopic.post', type='GROUP_THEME', access_token=Token)

                    signature_for_post = hashlib.md5(
                        b"application_key=" + byte_application + b"attachment=" + byte_attachment
                        + b"format=jsongid=" + byte_group_id +
                        b"hidden_post=truemethod=mediatopic.post"
                        b"type=GROUP_THEME" + byte_session).hexdigest()

                print(str(signature_for_post))

                data['sig'] = str(signature_for_post)
                url = 'https://api.ok.ru/fb.do'

                response = requests.post(url, data)
                print(response.json())
                posts.append('https://ok.ru/group/' + str(group_id) + '/topic/' + str(response.json()))

        return posts

class Gui(Toplevel, Vk, Ok):
    def __init__(self, parent, title="Обработка файлов"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def search_folder_for_excel_file(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def get_columns_data(self):
        # Получим столбцы из эксель файла
        try:
            wb = openpyxl.load_workbook(str(self.text_1.get()))
        except FileNotFoundError:
            message = 'Укажите путь к таблице!'
            ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
        sheet = wb.worksheets[0]
        private_post = []
        message = []
        path_to_file = []
        link_button = []
        link_title = []
        link_image = []
        post_with_button = []
        type_of_attachment = []
        name_of_video = []
        for i in range(1, sheet.max_row):
            if ((sheet.cell(row=i, column=1).value) == None):
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(2, max_row + 1):
            private_post.append(sheet.cell(row=i, column=1).value)
            post_with_button.append(sheet.cell(row=i, column=2).value)
            message.append(sheet.cell(row=i, column=3).value)
            type_of_attachment.append(sheet.cell(row=i, column=4).value)
            name_of_video.append(sheet.cell(row=i, column=5).value)
            path_to_file.append(str(sheet.cell(row=i, column=6).value).replace('\\','/'))
            link_button.append(sheet.cell(row=i, column=7).value)
            link_title.append(sheet.cell(row=i, column=8).value)
            link_image.append(sheet.cell(row=i, column=9).value)
        return private_post,post_with_button, message,type_of_attachment, \
               path_to_file,link_button,link_title,link_image,name_of_video

    def refill_table(self,posts):
        wb = openpyxl.load_workbook(str(self.text_1.get()))
        sheet = wb.worksheets[0]
        for i in range(0, len(posts)):
            print(posts[i])
            sheet.cell(row=i+2, column=10).value = str(posts[i])
        wb.save(str(self.text_1.get()))
        message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
        print('ok')
        return {}

    def start(self):
        private_post, post_with_button, message, type_of_attachment, \
        path_to_file, link_button, link_title, link_image, name_of_video = self.get_columns_data()
        Token = str(self.text_3.get())
        group_id = str(self.text_2.get())
        if self.var_1.get() and not self.var_2.get():
            print('vk')
            posts = self.start_Vk(private_post, post_with_button, message, type_of_attachment,
                                  path_to_file, link_button, link_title, link_image, name_of_video,
                                  Token,group_id)
        elif self.var_2.get() and not self.var_1.get():
            print('ok')
            posts = self.start_ok(private_post, post_with_button, message, type_of_attachment,
                                  path_to_file, link_button, link_title, link_image, name_of_video)
        elif self.var_1.get() and self.var_2.get():
            print('vk and ok')
        else:
            message = 'Укажите требуемую социальную сеть'
            ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
            print('lax')
            return {}
        return self.refill_table(posts)

    def text_4_on(self):
        if self.var_2.get():
            self.text_4["state"] = "normal"
            self.text_4.delete(0, END)
            self.text_4.insert(END, "Укажите Сессионный ключ")
        else:
            self.text_4["state"] = "disabled"

    def dialog(self, parent):
        self.parent = parent

        # Created main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_excel_file)

        self.label_2 = Label(parent, text="Укажите id сообщества")
        self.text_2 = Entry(parent)

        self.label_3 = Label(parent, text="Укажите Токен")
        self.text_3 = Entry(parent)

        self.var_1 = IntVar()
        self.var_2 = IntVar()

        self.chk_1 = Checkbutton(parent, text="VK", variable=self.var_1)

        self.chk_2 = Checkbutton(parent, text="ОК", variable=self.var_2, command=self.text_4_on)
        self.text_4 = Entry(parent, width=30, state=DISABLED, disabledforeground=parent.cget('bg'))


        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        self.label_2.pack()
        self.text_2.pack()

        self.label_3.pack()
        self.text_3.pack()

        self.chk_1.pack()
        self.chk_2.pack()
        self.text_4.pack()

        # start button
        self.but_start = Button(parent, text="Выполнить", command=self.start)
        self.but_start.pack()




if __name__ == "__main__":
        root = Tk()
        root.minsize(width=500, height=400)
        gui = Gui(root)
        root.mainloop()
